#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   BaseAlphaV2.py
@Time    :   2022/08/14 21:28:20
@Author  :   Liu Junli
@Version :   1.0
@Contact :   liujunli.xmu@foxmail.com
'''

import os
import datetime
import pandas as pd
from typing import Dict, Tuple, List, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import multiprocessing as mp

import utils
from serial import Serializable
from logger import get_module_logger


class BaseDataSet(Serializable):
    def __init__(self, provider, *args, **kwargs) -> None:
        super().__init__()
        self._provider = provider(*args, **kwargs)

        self.base_data = pd.DataFrame()  # 日频数据,准备其他数据的来源, cols = ['close', 'ind', 'cap', 'size', 'weight']
        self.style_factor = pd.DataFrame()  # barra因子 cols = C.cne5
        self.ind_dummy = pd.DataFrame()  # 行业dummy变量
        self.fields = pd.DataFrame()  # 要测试因子的域: index应该和base_data保持一致,value为0,1
        self.inds_name_code: Dict = {}  # 行业名称和代码对应关系

        self.prepare_data()

    def prepare_data(self):
        # 在该方法中准备好以上的五个dataframe
        # load data from database
        assert self._provider is not None, 'provider is None'
        with self._provider.start():
            ...
        ...
        raise NotImplementedError


class BaseAlphaSet(Serializable):
    logger = get_module_logger('AlphaSet')

    def __init__(self):
        super().__init__()
        self._alpha = pd.DataFrame()
        # !index.names=['tradedate','wind_code']
        # !'tradedate'应该是datetime.date格式, pd.to_datetime(yourcols).dt.date转换即可
        # example:
        # *factor.index.names=['tradedate,'wind_code']
        # *factor.reset_index(inplace=True)
        # *factor['%date'] = pd.to_datetime(factor['%date']).dt.date
        # *factor.set_index(['tradedate','wind_code'])

    def load_data(self):
        # load data from database
        self._provider = ...
        assert self._provider is not None, 'provider is None'
        with self._provider.start():
            self.alpha = self._provider.read()
            ...
        ...
        raise NotImplementedError

    @property
    def alpha(self):
        return self._alpha

    @alpha.setter
    def alpha(self, df: pd.DataFrame):
        assert df.index.names == ['tradedate', 'wind_code'], 'index.names须设置为["tradedate","wind_code"]'
        if not isinstance(df.index.levels[0][0], datetime.date):
            self.logger.warning('tradedate不是datetime.date格式! 尝试转换中...')
            try:
                df = df.reset_index()
                df['tradedate'] = pd.to_datetime(df['tradedate']).dt.date
                df = df.set_index(['tradedate', 'wind_code'])
            except Exception as e:
                self.logger.error('尝试转换失败, 请检查日期格式!', e)
                raise e
            self.logger.info('转换成功!')
        df.columns = ['alpha_' + col + '_raw' if not col.startswith('alpha_') else col + '_raw' for col in df.columns]
        self._alpha = df

    def pre_process(self,
                    dataset: BaseDataSet,
                    mode: str = 'mad',
                    multiple: int = 5,
                    method: str = 'ffill',
                    ifmp: bool = False,
                    w_method: str = 'cap',
                    freq: Optional[str] = None,
                    iforth: bool = True,
                    iffill: bool = False,
                    suffix_old: str = 'raw'):

        if freq:
            self.alpha = utils.resample(self.alpha, freq)

        x: pd.DataFrame
        n: int  # 进程数量

        if ifmp:
            pool = mp.Pool(processes=n)
            runs = []
            runs.append(
                pool.apply_async(utils.pre_process,
                                 args=(self.alpha, dataset.base_data, x, w_method, 'cne5', mode, multiple, method, iforth, iffill, suffix_old)))
            runs.append(...)
            result = [r.get() for r in runs]
            pool.close()
            pool.join()
        else:
            result = []
            result.append(utils.pre_process(self.alpha, dataset.base_data, x, w_method, 'cne5', mode, multiple, method, iforth, iffill, suffix_old))
            result.append(...)
        self.alpha = pd.concat([self.alpha] + result, axis=1)

        raise NotImplementedError


class BaseFactorCheck:
    def __init__(self, dataset: BaseDataSet, alphaset: BaseAlphaSet) -> None:
        self.dataset = dataset
        self.factor_data = utils.get_factor_data(alphaset.alpha, dataset.base_data)
        self.alpha_names = utils.get_alpha_names(self.factor_data.columns)

        # 回测结果
        self.IC = {}
        self.group_return = {}
        self.long_short = {}
        self.factor_return = {}

        # 回测信息
        self.factor_check_info = {}
        self.check_info()

        # 单因子回测结果：
        self.image = {}
        self.table = {}
        self.data = {}

    def check_info(self):
        self.factor_check_info = {
            '因子名称': None,
            '全市场域': self.dataset.fields.index[self.dataset.fields['all_A']].get_level_values(1).nunique(),
            '自定义域': self.dataset.fields.index[self.dataset.fields['user_defined']].get_level_values(1).nunique(),
            '回测起始': self.factor_data.index.levels[0][0].strftime('%Y-%m-%d'),
            '回测截至': self.factor_data.index.levels[0][-1].strftime('%Y-%m-%d'),
            '回测因子': None,
            '对照因子': None,
            '回测耗时': None,
        }

    @utils.func_timer
    def run(self):
        self.IC_analysis()
        self.factor_group_analysis()
        self.long_short_analysis()
        self.regression_analysis()
        [self.factor_report(alpha_name) for alpha_name in self.alpha_names]

    def factor_report(self, alpha_name):
        self.get_table(alpha_name)
        self.plot(alpha_name)
        self.render(alpha_name)

    def get_table(self, alpha_name):
        self.factor_check_info['因子名称'] = alpha_name
        self.table.update({'info': pd.DataFrame(self.factor_check_info, index=['取值']).T})

        cols: List[str]  # = func(alpha_name, self.factor_data.columns), 其中: func: Callable #*Note 以下三条语句的cols可能是不同的.
        reg: str  # = one of regs (self.factor_return.keys())
        ic: str  # = one of ICs (self.IC.keys())
        group: str  # = one of groups (self.group_return.keys())
        self.table.update({'stats': utils.factor_return_stats(self.factor_data[cols], self.factor_return[reg][cols])})
        self.table.update({'IC': utils.print_IC_table(self.IC[ic][cols])})
        self.table.update({'groups': utils.group_return_stats(self.group_return[group][cols])})
        ...
        return NotImplementedError

    def plot(self, alpha_name: str):

        self.image.clear()
        self.data.clear()

        name_show: str
        cols: List[str]  # = func(alpha_name, self.factor_data.columns)  #*Note 以下语句的cols可能是不同的.
        reg: str  # = one of regs (self.factor_return.keys())
        ic: str  # = one of ICs (self.IC.keys())
        group: str  # = one of groups (self.group_return.keys())

        # image1
        data = self.IC[ic][cols]
        suptitle = '全市场股票池IC累计值'
        self.image.update(dict(IC=utils.plot_line(data, ylabel='IC累计值', method='sum', suptitle=suptitle, alpha_name=alpha_name)))
        self.data.update({suptitle: data})
        # image2
        data = self.factor_return[reg].xs('beta', level=1).loc[:, cols]
        suptitle = '回归系数累计值'
        self.image.update(dict(reg=utils.plot_line(data, test='fmtest', suptitle=suptitle, alpha_name=alpha_name)))
        self.data.update({suptitle: data})
        # image3
        data = self.group_return[group].loc[:, cols].squeeze().unstack(level='fac_qt')
        suptitle = '%s行业分层分组测试' % name_show
        self.image.update(dict(groups2=utils.plot_line(data, test='mttest', suptitle=suptitle, alpha_name=alpha_name)))
        self.data.update({suptitle: data})
        ...
        return NotImplementedError

    def render(self,
               alpha_name: str,
               template_path: str = 'template.xlsx',
               save_path: str = 'factor_report',
               imgsize: Tuple = (602.0016, 287.3881)) -> None:

        out_path = '{}.xlsx'.format(alpha_name)
        if save_path is not None:
            out_path = os.path.join(save_path, out_path)
            if not os.path.exists(save_path):
                os.makedirs(save_path)
        wb = openpyxl.load_workbook(template_path)
        tab_pos, img_pos = utils.get_pos(wb)
        for img_name, img in self.image.items():
            sheet_name, row, col = img_pos[img_name]
            img = Image(img)
            img.width, img.height = imgsize
            wb[sheet_name].add_image(img, anchor=get_column_letter(col) + str(row))
        for tab_name, tab in self.table.items():
            sheet_name, row, col = tab_pos[tab_name]
            utils.excel_writer(tab, wb[sheet_name], row, col)
        for img_name, idata in self.data.items():
            utils.excel_writer(idata, wb[img_name])
        wb.save(out_path)

    @utils.func_timer
    def IC_analysis(self):
        # for example
        ICs: list = ['ic1', 'ic2', 'ic3']
        self.IC.update(ic1=utils.factor_information_coefficient(self.factor_data))
        self.IC.update(ic2=utils.factor_information_coefficient(self.factor_data, method='pearson'))
        self.IC.update(ic2=utils.factor_information_coefficient(self.factor_data, method='pearson', detail=True))
        ...
        return NotImplementedError

    @utils.func_timer
    def factor_group_analysis(self):
        # for example
        alpha_cols: list = ['alpha1', 'alpha2', ...]
        self.group_return.update(dict(group1=utils.group_return(self.factor_data[alpha_cols])))
        self.group_return.update(dict(group2=utils.group_return(self.factor_data[alpha_cols])))
        ...
        return NotImplementedError

    @utils.func_timer
    def long_short_analysis(self):
        # fro example
        groups1: list = ['group1', 'group2', ...]  # 在factor_group_analysis里的分组
        groups2: list = ['group3', 'group4', ...]  # 在factor_group_analysis里的分组
        self.long_short.update(dict(long_short1=pd.concat([utils.long_short_return(self.group_return(g), g) for g in groups1], axis=1)))
        self.long_short.update(dict(long_short2=pd.concat([utils.long_short_return(self.group_return(g), g) for g in groups2], axis=1)))
        ...
        return NotImplementedError

    @utils.func_timer
    def regression_analysis(self):
        regs: list = ['reg1', 'reg2', ...]
        self.factor_return.update(reg1=utils.factor_return(self.factor_data))
        return NotImplementedError
