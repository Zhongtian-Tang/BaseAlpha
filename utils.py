'''
@File    :   utils.py
@Time    :   2023/07/27 13:10:00
@Author  :   ZHongtian Tang
@Version :   1.0
@Contact :   799138793@qq.com
'''
import os
import io
import re
import time
import bisect
import multiprocessing as mp
import numpy as np
import pandas as pd
import xarray as xr
import statsmodels.api as sm
from scipy import stats
from typing import List, Union
from functools import wraps
import matplotlib.pyplot as plt
import openpyxl
from tqdm import tqdm
from openpyxl.worksheet.worksheet import Worksheet

tqdm.pandas(desc='process bar')

plt.rcParams['font.sans-serif'] = ['SimHei']  #用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  #用来正常显示负号

idx = pd.IndexSlice


def func_timer(function):
    @wraps(function)
    def function_timer(*args, **kwargs):
        print('[Function: {name} start...]'.format(name=function.__name__))
        t0 = time.time()
        result = function(*args, **kwargs)
        t1 = time.time()
        print('[Function: {name} finished, spent time: {time:.2f}s]'.format(name=function.__name__, time=t1 - t0))
        return result

    return function_timer


@func_timer
def factor_information_coefficient(factor_data: pd.DataFrame, detail: bool = False, method: str = 'spearman'):
    def pd_ic(factor_data_: pd.DataFrame, detail: bool = detail):
        alpha_col = factor_data_.columns.str.startswith('alpha')
        alpha = factor_data_.loc[:, alpha_col]
        ret = factor_data_.loc[:, 'Return']
        if not detail:
            return alpha.corrwith(ret, method=method)
        _ic = alpha.corrwith(ret, method='pearson')
        _ic_rank = alpha.corrwith(ret, method='spearman')
        return pd.DataFrame([_ic, _ic_rank], index=['IC', 'IC_rank'])

    cols = list(factor_data.columns[factor_data.columns.str.startswith('alpha')])
    cols.append('Return')
    return factor_data.groupby(['tradedate'])[cols].apply(pd_ic)


@func_timer
def group_return(factor_data: pd.DataFrame,
                 group_adjust: Union[List, None] = None,
                 out_group: Union[List, None] = None,
                 freq: Union[str, None] = None,
                 n_groups: int = 5,
                 w_method: str = 'equal'):

    alpha_cols_sel = factor_data.columns.str.startswith('alpha')
    if alpha_cols_sel.sum() > 1:
        alpha_cols = list(factor_data.columns[alpha_cols_sel])
        other_cols = list(factor_data.columns[~alpha_cols_sel])
        pool = mp.Pool(processes=mp.cpu_count())
        out_ = [
            pool.apply_async(mean_return_by_quantile, args=(factor_data[other_cols + [alpha]], group_adjust, n_groups, w_method, out_group, freq))
            for alpha in alpha_cols
        ]
        out_list = [o.get() for o in out_]
        pool.close()
        pool.join()
        out_list.sort(key=len, reverse=True)
        out = pd.concat(out_list, axis=1).sort_index(axis=0)
        if not out.index.names:
            try:
                out.index.names = out_list[0].index.names
            except:
                raise ValueError('所有因子分组结果均为空！')
    else:
        out = mean_return_by_quantile(factor_data, group_adjust, n_groups, w_method, out_group, freq)
    out = out.reindex(
        pd.MultiIndex.from_product([out.index.levels[i] for i in range(out.index.nlevels)],
                                   names=[out.index.names[i] for i in range(out.index.nlevels)]))
    return out


@func_timer
def mean_return_by_quantile(factor_data: pd.DataFrame,
                            group_adjust: Union[List, None] = None,
                            n_groups: int = 5,
                            w_method: str = 'equal',
                            out_group: Union[List, None] = None,
                            freq: Union[str, None] = None):
    """
    group_adjust: ['ind','size'] or ['ind'] or ['size'] or None
    out_group: ['ind','size'] or ['ind'] or ['size'] or None
    """
    grouper = ['tradedate']

    if group_adjust:
        grouper.extend(group_adjust)

    def df_qcut(df: pd.DataFrame, n_groups: int = n_groups):
        if df.empty or df.count().sum() == 0:
            return pd.DataFrame().reindex_like(df)

        def q_cut(x: pd.Series, n_groups: int = n_groups):
            try:
                return pd.qcut(x, q=n_groups, duplicates='drop', labels=['p' + str(i) for i in range(1, n_groups + 1)])
            except:
                try:
                    return pd.qcut(x.rank(method='first'),
                                   q=np.linspace(0, 1, num=n_groups + 1),
                                   duplicates='drop',
                                   labels=['p' + str(i) for i in range(1, n_groups + 1)])
                except:
                    try:
                        return pd.qcut(x.rank(method='first'), q=n_groups, duplicates='drop', labels=['p' + str(i) for i in range(1, n_groups + 1)])
                    except ValueError as e:
                        _, n_bins = pd.qcut(x.rank(method='first'), q=n_groups, duplicates='drop', retbins=True, labels=False)
                        if len(n_bins) > 0:
                            return pd.qcut(x.rank(method='first'),
                                           q=n_groups,
                                           duplicates='drop',
                                           labels=['p' + str(i) for i in range(1, len(n_bins))])
                        else:
                            print(x)
                            raise ValueError(len(x.unique()), repr(e))

        return df.apply(q_cut)

    col = factor_data.columns[factor_data.columns.str.startswith('alpha')].values
    if freq is not None:
        temp = factor_data.swaplevel(axis=0).sort_index(axis=0)
        temp = resample(temp, freq).groupby(grouper)[col].apply(df_qcut, n_groups=n_groups).reindex(temp.index, method='ffill')
        factor_data['fac_qt'] = temp.swaplevel(axis=0).sort_index(axis=0)
    else:
        factor_data['fac_qt'] = factor_data.groupby(grouper)[col].apply(df_qcut, n_groups=n_groups)

    def weight_func(df: pd.DataFrame, w_method: str = w_method):
        if w_method == 'equal':
            return df['Return'].mean()
        elif w_method == 'cap':
            return df['Return'].mul((df['cap'] / df['cap'].sum())).sum()

    out_grouper = ['tradedate', 'fac_qt']
    if out_group:
        out_grouper.extend(out_group)

    out = factor_data.groupby(out_grouper)[['Return', 'cap']].apply(weight_func)
    if out.empty:
        # factor_data[col[0]] = np.random.random(factor_data.shape[0])
        # factor_data['fac_qt'] = factor_data.groupby(grouper)[col].apply(df_qcut)
        # out = factor_data.groupby(out_grouper)[['Return', 'cap']].apply(lambda x: np.nan)
        return pd.Series(name=col[0])
    return out.to_frame(col[0])


@func_timer
def factor_return1(factor_data: pd.DataFrame, detail: bool = False) -> pd.DataFrame:
    def reg_wls(df: pd.DataFrame, detail: bool = detail):
        y = df['Return']
        w = df['weight']
        alpha_col = df.columns[df.columns.str.startswith('alpha')]
        x = sm.add_constant(df[alpha_col])
        index_ = set(y.dropna(how='any').index) & set(x.dropna(how='any').index) & set(w.dropna(how='any').index)
        fit_ = sm.WLS(y[index_], x.loc[index_], weights=w[index_]).fit()
        if not detail:
            return pd.DataFrame([fit_.params[1]], columns=alpha_col, index=['beta'])
        return pd.DataFrame([fit_.params[1], fit_.tvalues[1]], columns=alpha_col, index=['beta', 'tvalues'])

    cols = list(factor_data.columns[factor_data.columns.str.startswith('alpha')])
    ols_single = lambda col: factor_data.groupby(['tradedate'])[['Return', 'weight'] + [col]].apply(reg_wls)
    return pd.concat([ols_single(col) for col in cols], axis=1)


@func_timer
def pre_process(alpha: pd.DataFrame,
                base_data: pd.DataFrame,
                x: Union[pd.DataFrame, None] = None,
                w_method: str = 'cap',
                suffix_new: str = 'new',
                mode: str = 'mad',
                multiple: int = 5,
                method: str = 'ffill',
                iforth: bool = True,
                iffill: bool = False,
                suffix_old: str = 'raw') -> None:

    w = base_data['weight'].to_frame()
    cap = base_data['cap'].to_frame()

    alpha = drop_extremevalue(alpha, mode=mode, multiple=multiple)
    if iffill:
        alpha = process_nan(alpha, method=method, ind=base_data['ind'].to_frame())
    # alpha = alpha.fillna(0)
    if iforth:
        try:
            alpha = orth(alpha, x, w)
        except:
            try:
                alpha = orth2(alpha, x, w)
            except:
                alpha = orth1(alpha, x, w)
    alpha = zscore(alpha, cap, w_method=w_method)
    if isinstance(alpha, pd.Series):
        alpha = pd.DataFrame(alpha)
    alpha.columns = alpha.columns.str.replace('_' + suffix_old, '_' + suffix_new)
    return alpha


def process_nan(alpha: pd.DataFrame, method: str = 'ffill', ind: Union[pd.DataFrame, None] = None) -> pd.DataFrame:
    if method == 'ffill':
        alpha = alpha.groupby('wind_code').apply(lambda x: x.fillna(method='ffill', axis=0))
    if method == 'mean':
        alpha = alpha.groupby('tradadate').apply(lambda x: x.fillna(value=x.mean(), axis=0))
    if method == 'indmean':
        alpha = alpha.join(ind, how='left').groupby(['tradedate', 'ind']).apply(lambda x: x.fillna(value=x.mean(), axis=0)).drop(columns='ind')
        #有可能整个行业都是NaN,所以按行业均值填充完还需要按全部均值再填充一次
    return alpha.fillna(value=alpha.mean())


def drop_extremevalue(alpha: pd.DataFrame, mode: str = 'mad', multiple: int = 5) -> pd.DataFrame:
    def drop_ev(alpha: pd.DataFrame, mode: str, multiple: str):
        if mode == 'mad':
            med = alpha.median()
            center = med
            abs_med = (alpha - med).abs().replace(np.inf, np.nan)
            diff = multiple * abs_med.median()
        elif mode == 'std':
            x = alpha.mask(np.isinf(alpha))
            center = x.mean(axis=0)
            diff = multiple * x.std(ddof=0)
        return alpha.clip(lower=center - diff, upper=center + diff, axis=1)

    return alpha.groupby('tradedate').apply(drop_ev, mode, multiple)


@func_timer
def orth2(alpha: pd.DataFrame, x: pd.DataFrame, w: pd.DataFrame):
    def _ols(df: pd.DataFrame):
        alpha_num = df.columns.str.startswith('alpha').sum()
        alpha = df.iloc[:, :alpha_num]
        x = df.iloc[:, alpha_num:-1]
        w = df.iloc[:, -1]

        def ols_main(ialpha: pd.DataFrame, x: pd.DataFrame = x, w: pd.DataFrame = w):
            try:
                return sm.WLS(ialpha, x, weights=w, missing='drop').fit().resid.reindex(ialpha.index)
            except:
                return pd.Series(np.nan, index=ialpha.index)

        return pd.concat([ols_main(alpha.iloc[:, i]) for i in range(alpha_num)], axis=1)

    out = pd.concat([alpha, sm.add_constant(x), w], axis=1).reindex(alpha.index).groupby('tradedate').apply(_ols)
    out.columns = alpha.columns
    return out


@func_timer
def orth1(alpha: pd.DataFrame, x: pd.DataFrame, w: pd.DataFrame) -> pd.DataFrame:

    n = 200
    ndate = len(alpha.index.levels[0])
    get_idx = lambda i: idx[alpha.index.levels[0][i * n:np.min([(i + 1) * n, ndate])], :]
    cores = mp.cpu_count()
    pool = mp.Pool(processes=cores)
    out_ = [
        pool.apply_async(orth2, args=(alpha.loc[get_idx(i), :], x.loc[get_idx(i), :], w.loc[get_idx(i), :])) for i in range(int(np.ceil(ndate / n)))
    ]
    out_list = [o.get() for o in out_]
    pool.close()
    pool.join()

    return pd.concat(out_list, axis=0).sort_index(axis=0)


def zscore(alpha: pd.DataFrame, cap: Union[pd.DataFrame, None] = None, w_method: str = 'equal') -> pd.DataFrame:
    if w_method == 'equal':
        return alpha.groupby('tradedate').apply(lambda x: ((x - x.mean(axis=0)) / x.std(axis=0).replace(0, np.nan)))
    elif w_method == 'cap':
        fac_data = alpha.join(cap, how='left')
        _zscore = lambda df: (df - df.mul(df['cap'] / df['cap'].sum(), axis=0).sum(axis=0)) / df.std(axis=0).replace(0, np.nan)
        return fac_data.groupby('tradedate').apply(_zscore)[alpha.columns]


def print_IC_table(data: Union[pd.DataFrame, pd.Series]) -> pd.DataFrame:
    def calc_stats(data: pd.DataFrame) -> pd.DataFrame:
        data_stats = pd.DataFrame()
        data_stats['IC_obs'] = data.count()
        data_stats['IC_mean'] = data.mean()
        data_stats['IC_std'] = data.std()
        data_stats['IC_IR'] = data_stats['IC_mean'] / data_stats['IC_std']
        data_stats['|IC|>0.02'] = data.apply(lambda x: x.abs() > 0.02).sum() / data_stats['IC_obs']
        data_stats.index.name = 'factor'
        return data_stats

    if isinstance(data, pd.Series):
        data = pd.DataFrame(data)

    try:
        data.index = pd.to_datetime(data.index)
    except:
        raise ValueError('index不能转换为日期格式')

    try:
        data.columns = ['fac_' + col.split('_')[-1] if 'alpha' in col else col for col in data.columns]
    except:
        pass

    stats_year = data.groupby(data.index.year).apply(calc_stats)

    stats_all = calc_stats(data)
    stats_all['tradedate'] = '统计以来'
    stats_all = stats_all.reset_index().set_index(['tradedate', 'factor'])

    return pd.concat([stats_year, stats_all])


def excel_writer1(data: Union[pd.DataFrame, pd.Series], sheet, header: bool = True, index: bool = True, startrow=0, startcol=0):

    if isinstance(data, pd.DataFrame):
        data = pd.DataFrame(data)
    if index:
        data = data.reset_index()
    if header:
        sheet.write_row(startrow, startcol, data.columns) if isinstance(startcol, int) else sheet.write_row(startcol + str(startrow), data.columns)
        startrow += 1
    [
        sheet.write_row(startrow + i, startcol, data.iloc[i, :].values) if isinstance(startcol, int) else sheet.write_row(
            startcol + str(startrow + i), data.iloc[i, :].values) for i in range(len(data))
    ]


def factor_return_stats(fac_data: pd.DataFrame, fac_ret: pd.DataFrame) -> pd.DataFrame:

    out = pd.DataFrame()
    out['fac_mean'] = fac_data.mean(axis=0)
    out['fac_std'] = fac_data.std(axis=0)
    out['fac_skew'] = fac_data.skew(axis=0)
    out['Yr_ICIR'] = fac_ret.mean(axis=0) / fac_ret.std(axis=0) * np.sqrt(252)
    out['ret_Yr'] = (fac_ret + 1).prod(axis=0)**(252 / len(fac_ret)) - 1
    ttest = lambda x: stats.ttest_1samp(x.dropna(), 0).statistic
    out['FM_test'] = fac_ret.apply(ttest)
    try:
        out.index = ['fac_' + index.split('_')[-1] if '_' in index else index for index in out.index]
    except:
        pass
    return out


def group_return_stats(group_ret: pd.Series) -> pd.DataFrame:

    out = pd.DataFrame()
    gret_gb = group_ret.groupby('fac_qt')
    out['alpha_Yr'] = gret_gb.apply(lambda x: (x + 1).prod()**(252 / len(x)) - 1)
    out['alpha_IR'] = gret_gb.mean().div(gret_gb.std()) * np.sqrt(252)

    return out


def plot_line(data: Union[pd.DataFrame, pd.Series],
              test: str = '',
              method: str = 'prod',
              ylabel: str = '累计收益',
              suptitle: str = '',
              title: str = '',
              leg_title=None,
              leg_txt=None,
              return_fig: bool = False,
              save_local: bool = False,
              alpha_name: Union[str, None] = None,
              path: str = 'image'):

    if isinstance(data, pd.Series):
        data = pd.DataFrame(data)

    if not isinstance(data.index, pd.DatetimeIndex):
        data.index = pd.to_datetime(data.index)

    try:
        data.columns = ['fac_' + col.split('_')[-1] if 'alpha' in col else col for col in data.columns]
    except:
        pass

    if test == 'fmtest':
        title = 'FM检验: ' + fmtest2str(data)

    elif test == 'mttest':
        title = '单调性: ' + mttest2str(data)

    if method == 'sum':
        data = data.cumsum(axis=0)

    elif method == 'prod':
        data = (data + 1).cumprod(axis=0)

    if not leg_txt:
        leg_txt = data.columns.to_list()

    fig = plt.figure(figsize=(8, 3.819095477386935))
    ax = fig.add_axes([0.1, 0.18, 0.88, 0.68])
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    ax.spines['left'].set_color('#b3b3cc')
    ax.spines['bottom'].set_color('#b3b3cc')
    ax.grid(b=None, which='major', axis='both', linestyle='--', alpha=0.3, linewidth=0.5)
    ax.plot(data, alpha=1, lw=1.5)
    ax.set_ylabel(ylabel, labelpad=5, fontsize=13)
    fig.suptitle(suptitle, fontsize=18, x=0.1, y=0.98, horizontalalignment='left')
    ax.set_title(title, fontsize=13, loc='left', pad=1)
    ax.legend(leg_txt,
              fontsize=14,
              title=leg_title,
              loc='lower center',
              bbox_to_anchor=(0.5, -0.1),
              borderaxespad=-1.5,
              frameon=False,
              ncol=len(data.columns))

    if return_fig:
        return fig

    if save_local:
        if alpha_name:
            suptitle += '_' + alpha_name

        if path:
            if not os.path.exists(path):
                os.makedirs(path)
            suptitle = os.path.join(path, suptitle)

        fig.savefig(suptitle + '.png', dpi=300)
        return suptitle + '.png'
    else:
        imgdata = io.BytesIO()
        fig.savefig(imgdata, format='png')
        return imgdata


def fmtest2str(data: pd.DataFrame) -> pd.DataFrame:
    return str(data.apply(lambda x: round(stats.ttest_1samp(x.dropna(), 0).statistic, 2)).to_dict()).replace('{', '').replace('}',
                                                                                                                              '').replace("'", '')


def mttest2str(data: pd.DataFrame) -> pd.DataFrame:
    grouper = [float(re.findall(r"\d+\.?\d*", col)[0]) for col in data.columns]
    return str(data.groupby(data.index.year).apply(lambda df: round(stats.pearsonr((df + 1).prod(axis=0), grouper)[0], 2)).to_dict()).replace(
        '{', '').replace('}', '')


@func_timer
def factor_return(factor_data: pd.DataFrame, detail: bool = False) -> pd.DataFrame:
    def main_ols(Return, alpha, weight, detail: bool = detail):
        if np.all(np.isnan(alpha)):
            if detail:
                return np.array([np.nan, np.nan])
            return np.array([np.nan])
        alpha = sm.add_constant(alpha)
        try:
            out = sm.WLS(Return, alpha, weights=weight, missing='drop').fit()
            if detail:
                return np.array([out.params[1], out.tvalues[1]])
            return np.array([out.params[1]])
        except:
            if detail:
                return np.array([np.nan, np.nan])
            return np.array([np.nan])

    def xr_ols(Return, alpha, weight, detail: bool = detail):
        results = xr.apply_ufunc(
            main_ols,
            Return,
            alpha,
            weight,
            input_core_dims=[['wind_code'], ['wind_code'], ['wind_code']],  # 每个参数只有一个条目的列表
            output_core_dims=[['results']],  # 计算beta:'variable',计算resid:'ticker'
            exclude_dims=set(('wind_code', )),  # 结果变量中要删除的变量
            kwargs={'detail': detail},  # !必须以关键字的形式传进去
            vectorize=True,
            dask="parallelized",
            output_dtypes=[alpha.dtype]  # 每个输出一个； 也可以是float或np.dtype（“ float64”）
        )
        results['results'] = ['beta', 'tvalue'] if detail else ['beta']
        return results

    alpha = factor_data[list(factor_data.columns[factor_data.columns.str.startswith('alpha')])].to_xarray().to_array(dim='factor')
    Return = factor_data['Return'].to_xarray()
    weight = factor_data['weight'].to_xarray()
    out = xr_ols(Return, alpha, weight, detail=detail).to_dataframe('').unstack(level=1)
    out.columns = out.columns.droplevel(0)
    return out


@func_timer
def orth(alpha: pd.DataFrame, x: pd.DataFrame, w: pd.DataFrame) -> pd.DataFrame:
    def main_ols(alpha, x, weight):
        resid = np.full_like(alpha, np.nan)
        flag = np.isnan(alpha) | np.isnan(x).sum(axis=1) | np.isnan(weight)
        x = sm.add_constant(x)
        try:
            resid[flag == 0] = sm.WLS(alpha, x, weights=weight, missing='drop').fit().resid
        except:
            pass
        return resid

    alpha_index = alpha.index
    alpha.columns.name = 'factor'
    alpha = alpha.stack().to_xarray()
    weight = w.reindex(alpha_index).squeeze().to_xarray()
    x = xr.Dataset.from_dataframe(x.reindex(alpha_index)).to_array(dim='x')
    wind_code = list(set(x['wind_code'].values) & set(alpha['wind_code'].values) & set(weight['wind_code'].values))
    tradedate = list(set(x['tradedate'].values) & set(alpha['tradedate'].values) & set(weight['tradedate'].values))
    alpha = alpha.reindex(wind_code=wind_code, tradedate=tradedate)
    x = x.reindex(wind_code=wind_code, tradedate=tradedate)
    weight = weight.reindex(wind_code=wind_code, tradedate=tradedate)

    resid = xr.apply_ufunc(main_ols,
                           alpha,
                           x,
                           weight,
                           input_core_dims=[['wind_code'], ['wind_code', 'x'], ['wind_code']],
                           output_core_dims=[['wind_code']],
                           vectorize=True,
                           dask="parallelized",
                           output_dtypes=[alpha.dtype]).to_dataframe('').unstack(level=1).reindex(alpha_index)
    resid.columns = resid.columns.droplevel(0)
    return resid


def read_h5(file_path: str, h5file, datakey, cond: str, columns=None) -> pd.DataFrame:

    if not datakey.startswith('/'): datakey = '/' + datakey
    fpath = os.path.join(file_path, h5file)
    with pd.HDFStore(fpath, mode="r") as hdf:
        qres = hdf.select(datakey, where=cond, columns=columns)
    return qres.reset_index()


def resample(data: pd.DataFrame, freq: str = 'M'):
    date_idx = pd.DatetimeIndex(data.index.get_level_values('tradedate').unique())
    year = date_idx.year
    if freq == 'D':
        freq = date_idx.dayofyear
    elif freq == 'Y':
        freq = date_idx.year
    elif freq == 'M':
        freq = date_idx.month
    elif freq == 'W':
        year = date_idx.isocalendar().year  #每年的最后一天不一定是当年（有可能是下一年）的星期，最开始一天不一定是当年（有可能是上一年）的星期
        freq = date_idx.isocalendar().week
    elif freq == 'Q':
        freq = date_idx.quarter
    date_idx = date_idx.to_frame().groupby([year, freq], as_index=False).last().set_index('tradedate')

    return data.reindex(date_idx.index, level='tradedate', method='ffill')



def resample1(data: pd.DataFrame, freq: str = 'M'):
    try:
        return data.groupby([pd.Grouper(freq=freq, level=0), data.index.get_level_values('wind_code')]).last()
    except:
        try:
            return data.unstack('wind_code').resample(freq).last().stack('wind_code')
        except:
            try:
                return data.reset_index(level='tradedate').groupby(
                    level='wind_code').apply(lambda x: x.set_index('tradedate').resample(freq).last()).swaplevel(axis=0).sort_index(axis=0)
            except:
                raise ValueError('resample error!')


def get_alpha_names(alpha_names: pd.Series):
    return alpha_names[alpha_names.str.startswith('alpha')].str.split('_', expand=True).levels[1]


def split_date(st, ed):
    if isinstance(st, str):
        st = st.replace('-', '')
    if isinstance(ed, str):
        ed = ed.replace('-', '')
    year = np.arange(int(st[:4]), int(ed[:4]) + 1, dtype=int)
    year_st = [str(i) + '0101' for i in year]
    year_ed = [str(i) + '1231' for i in year]
    year_st[0] = st
    year_ed[-1] = ed
    return year_st, year_ed


def excel_writer(data: Union[pd.DataFrame, pd.Series], ws: Worksheet, startrow=0, startcol=0, header: bool = True, index: bool = True):
    if isinstance(data, pd.DataFrame):
        data = pd.DataFrame(data)
    if index:
        data = data.reset_index()
    if header:
        data = data.T.reset_index().T

    for i, row in enumerate(data.itertuples(index=False)):
        for j, value in enumerate(row):
            ws.cell(row=startrow + i, column=startcol + j).value = value


def get_pos(wb: openpyxl.Workbook):
    table = {}
    image = {}
    for sn in wb.sheetnames:
        sh = wb[sn]
        for i in range(sh.min_row, sh.max_row + 1):
            for j in range(sh.min_column, sh.max_column + 1):
                if sh.cell(i, j).value is not None:
                    if str(sh.cell(i, j).value).startswith('# table'):
                        table.update({str(sh.cell(i, j).value).replace('# table:', ''): (sn, i, j)})
                    elif str(sh.cell(i, j).value).startswith('# image'):
                        image.update({str(sh.cell(i, j).value).replace('# image:', ''): (sn, i, j)})
    return table, image


def get_factor_data(alpha: pd.DataFrame, base_data: pd.DataFrame, lag: bool = False) -> pd.DataFrame:
    # alpha 中有可能有factor_data中没有的日期，直接用alpha的index给去factor reindex会导致部分日期return为0
    factor_data = alpha.join(base_data, how='left')  #!确保alpha.index是base_data.index的子集
    # 注意一定要sort_values()
    # _idx = sorted(set(base_data.index).union(set(alpha.index)))
    # factor_data = alpha.join(base_data.reindex(_idx).groupby('wind_code').fillna(method='ffill'), how='left')
    # factor_data = pd.concat([alpha, base_data], axis=1, join='outer').groupby('wind_code').fillna(method='ffill')
    # factor_data = factor_data.reindex(alpha.index)
    # index = base_data.index.append(alpha.index).drop_duplicates().sort_values()
    # factor_data = alpha.join(base_data.reindex(index).groupby('wind_code').fillna(method='ffill'), how='left')
    _ret_fn1 = lambda x: (x.shift(-1) / x.fillna(method='bfill')).shift(-1) - 1
    _ret_fn0 = lambda x: x.shift(-1) / x.fillna(method='bfill') - 1
    factor_data['Return'] = factor_data['close'].groupby('wind_code').apply(_ret_fn1 if lag else _ret_fn0)
    return factor_data


def handle_code(code: str) -> str:
    if code.startswith('9'):
        return '60' + code[1:] + '.SH'
    elif code.startswith('77'):
        return '300' + code[2:] + '.SZ'
    elif code.startswith('7'):
        return '00' + code[1:] + '.SZ'
    else:
        print('code: %s 转换失败' % code)


def long_short_return(data, group: str = 'group', grouer='tradedate') -> pd.DataFrame:
    def _long_short_return(df: pd.DataFrame, group=group):
        t, b = df.index.levels[1][[0, -1]].to_list()
        out = df.xs(b, level=1, axis=0) - df.xs(t, level=1, axis=0)
        out.columns = pd.MultiIndex.from_product([[group], out.columns], names=['field', 'factor'])
        return out

    return data.groupby(grouer).apply(_long_short_return).droplevel(level=0, axis=0)
