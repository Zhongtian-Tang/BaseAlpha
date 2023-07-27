#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   BaseAlpha.py
@Time    :   2022/07/15 14:36:34
@Author  :   Liu Junli
@Version :   1.0
@Contact :   liujunli.xmu@foxmail.com
'''

import os
import pickle
import warnings
import re
import numpy as np
import pandas as pd
import mysql.connector
import cx_Oracle
from sqlalchemy import create_engine
import xlsxwriter as xw
from typing import Callable, Dict, Tuple, Union, List
import multiprocessing as mp
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

import utils

warnings.filterwarnings("ignore")

BM = {
    'sz50': '000016.SH',
    'hs300': '000300.SH',
    'zz1000': '000852.SH',
    'zzlt': '000902.CSI',
    'zz500': '000905.SH',
    'zz800': '000906.SH',
    'zzqz': '000985.CSI'
}

settings = {
    "start_date": 20170101,  # 表示因子的开始结束时间，closeprice数据时间应该在end_date后再加一天
    "end_date": 20220630,
    "stock_pools": ['all_A', 'user_defined'],  # 'all_A', 'sz50', 'hs300', 'zz1000', 'zz800', 'zz500', 'zzqz'(中证全指), 'zzlt'(中证流通), 'user_defined';
    "style_factor":
    ['Beta', 'BooktoPrice', 'EarningsYield', 'Growth', 'Leverage', 'Liquidity', 'Momentum', 'NonLinearSize', 'ResidualVolatility', 'Size'],
    "alpha_factor_list": ['fearng'],  #['fearng'],  #['rec', 'sfy12p'],  #['Beta'],  #
    "data_sources": 'mysql',  # 'mysql'
    "alpha_factor_source":
    'EQU_FACTOR_AF',  #'equ_factor_af.h5',  #'EQU_FACTOR_AF',  #'tl_factor\equ_factor_af.h5',  #'barra_cne5_factor',  #'EQU_FACTOR_CF'
    "connect": dict(host='10.224.16.81', user='haquant', passwd='haquant', database='jydb', port=3306, charset='utf8'),
    "connect_or": dict(part1='rededata', part2='HAzc_1805@10.224.0.21:1521/datayesdb'),
    # 'rededata/HAzc_1805@10.224.0.21:1521/datayesdb',
    "file_path": r'D:\Data',  #r'\\10.224.1.70\public\data\dailydata_copy',
}

idx = pd.IndexSlice


class Connection:
    def __init__(self, settings) -> None:
        self.settings = settings
        self.state: bool = False
        if self.settings['data_sources'] == 'mysql':
            try:
                self.start()
            except:
                print('warning: mysql connection error!')

    def start(self) -> None:
        self.db = mysql.connector.connect(**self.settings['connect'])
        self.cr = self.db.cursor()
        self.conn = create_engine('oracle://' + self.settings['connect_or']['part1'] + ':' + self.settings['connect_or']['part2'])
        self.db_or = cx_Oracle.connect(self.settings['connect_or']['part1'] + '/' + self.settings['connect_or']['part2'])
        self.cr_or = self.db_or.cursor()
        self.state = True

    def delete(self, attr_to_del: Union[Tuple, None] = ('db', 'cr', 'conn', 'db_or', 'cr_or')) -> None:
        [delattr(self, iattr) if (iattr in self.__dict__.keys()) else None for iattr in attr_to_del]

    def close(self) -> None:
        if self.state:
            self.db.close()
            self.db_or.close()
            self.state = False


class Base(Connection):
    def __init__(self, settings) -> None:
        super().__init__(settings)

    @staticmethod
    def load(file_path: str = 'dataset.pkl'):
        return pickle.load(open(file_path, 'rb'))

    def save(self, file_path: str = 'dataset.pkl', data=None):
        if data is not None:
            pickle.dump(data, open(file_path, 'wb'))
        else:
            self.delete(('db', 'cr', 'conn', 'db_or', 'cr_or'))
            pickle.dump(self, open(file_path, 'wb'))

    def _load_data_from_sql(self, exp: str, con):
        path = os.path.join('temp', exp.replace(', ', ' ').replace(',', ' ').replace(' ', '_') + '.pkl')  #不要改
        if os.path.exists(path):
            data = self.load(path)
        else:
            if con == self.cr_or:
                self.cr_or.execute(exp)
                data = pd.DataFrame(self.cr_or.fetchall())
                try:
                    data.columns = [field.split('as')[-1].strip() for field in re.findall(r'SELECT(.*?)FROM', exp)[0].split(',')]
                except:
                    data.columns = ['tradedate', 'wind_code'] + list(self.settings['alpha_factor_list'])
            else:
                data = pd.read_sql(exp, con=con)

            if con == self.db:
                try:
                    data.set_index(['tradedate', 'wind_code'], inplace=True)
                except:
                    pass
            else:
                data['tradedate'] = pd.to_datetime(data['tradedate'], format='%Y%m%d').dt.date
                data['wind_code'] = data['wind_code'].str.replace('XSHE', 'SZ').str.replace('XSHG', 'SH').str.replace('XBEI', 'BJ')
                data.set_index(['tradedate', 'wind_code'], inplace=True)
            self.save(path, data)
        return data

    @utils.func_timer
    def read_sql(self, exp: str, con=None) -> pd.DataFrame:
        con = self.db if con is None else con
        exp = exp.strip().replace('\n', '')
        exp = ' '.join(exp.split())
        st = exp.split(' ')[-3]
        ed = exp.split(' ')[-1]
        st_, ed_ = utils.split_date(st, ed)
        data = []
        for ist, ied in zip(st_, ed_):
            iexp = exp.replace(st, ist).replace(ed, ied)
            data.append(self._load_data_from_sql(iexp, con))
        return pd.concat(data, axis=0, join='outer')

    def _load_data_from_h5(self, file_path, h5file, key, cond, columns=None):
        _col = [] if columns is None else columns
        file_name = h5file + '_' + key + '_' + '_'.join(_col) + '_' + cond.replace('>=', 'Ge').replace('<=', 'Le') + '.pkl'
        path = os.path.join('temp', file_name)
        if os.path.exists(path):
            data = self.load(path)
        else:
            data: pd.DataFrame = pd.read_hdf(os.path.join(file_path, h5file), key=key, where=cond, columns=columns, mode='r').reset_index()
            try:
                data = data.set_index(['tradedate', 'wind_code']).sort_index(axis=1)
            except:
                pass
            self.save(path, data)
        return data

    @utils.func_timer
    def read_h5(self, h5file: str, key: str, file_path: str = None, cond: str = None, columns: list = None):
        file_path = self.settings['file_path'] if not file_path else file_path
        cond = "tradedate>='%s'&tradedate<='%s'" % (str(self.settings['start_date']),
                                                    str(self.settings['end_date'])) if not cond else cond.strip().replace('\n', '')
        cond = ' '.join(cond.split())
        st = cond.replace('\n', '').split("'")[-4]
        ed = cond.replace('\n', '').split("'")[-2]
        st_, ed_ = utils.split_date(st, ed)
        data = []
        for ist, ied in zip(st_, ed_):
            icond = cond.replace(st, ist).replace(ed, ied)
            data.append(self._load_data_from_h5(file_path, h5file, key, icond, columns))
        return pd.concat(data, axis=0, join='outer')


class BaseDataSet(Base):
    def __init__(self, settings) -> None:
        # 索引都是两级：tradedate, wind_code
        super().__init__(settings)
        self.base_data = pd.DataFrame()  # 日频数据,准备其他数据的来源
        self.style_factor = pd.DataFrame()  # barra因子
        self.ind_dummy = pd.DataFrame()  # 行业dummy变量
        self.fields = pd.DataFrame()  # 要测试因子的域: 权重为0,1

        self.inds_name_code: Dict = {}
        self.bm_name_code: Dict = BM

        self.prepare_data()  # 读取并准备数据
        self.close()

    @utils.func_timer
    def load_data_sql(self):

        load_data_col = ['closeprice', 'firstindustrycode', 'afloats', 'adjclose']
        loaddata_sqlexp = """
        SELECT tradedate, wind_code, %s FROM dailydata WHERE tradedate BETWEEN %s AND %s
        """ % (', '.join(load_data_col), str(self.settings['start_date']), str(self.settings['end_date']))
        self.load_data = self.read_sql(loaddata_sqlexp)

        ind_sql = """
        SELECT DISTINCT firstindustryname as ind_name, firstindustrycode as ind_code FROM dailydata WHERE tradedate='2020-12-31'
        """
        self.inds_name_code = pd.read_sql(ind_sql, con=self.db).dropna(how='any').astype({
            'ind_code': int
        }).set_index('ind_code')['ind_name'].to_dict()

        barrafactor_sqlexp = """
        SELECT tradedate, wind_code, %s FROM barra_cne5_factor
        WHERE tradedate BETWEEN %s AND %s
        """ % (', '.join(self.settings['style_factor']), str(self.settings['start_date']), str(self.settings['end_date']))
        self.style_factor = self.read_sql(barrafactor_sqlexp)

        self.fields = pd.concat([self.prepare_fields(stock_pool) for stock_pool in self.settings['stock_pools']], axis=1)

    @utils.func_timer
    def load_data_h5(self):
        load_data_col = ['closeprice', 'firstindustrycode', 'firstindustryname', 'afloats', 'adjclose']
        self.load_data = self.read_h5(h5file='stock_daily.h5', key='dailydata', columns=load_data_col)
        self.style_factor = self.read_h5(h5file='barra_cne5_factor_value.h5', key='factor_value', columns=self.settings['style_factor'])
        self.inds_name_code = self.load_data[['firstindustrycode',
                                              'firstindustryname']].drop_duplicates().dropna(how='any').rename(columns={
                                                  'firstindustrycode': 'ind_code',
                                                  'firstindustryname': 'ind_name'
                                              }).astype({
                                                  'ind_code': int
                                              }).set_index('ind_code')['ind_name'].to_dict()
        self.fields = pd.concat([self.prepare_fields(stock_pool) for stock_pool in self.settings['stock_pools']], axis=1, join='outer')

    @utils.func_timer
    def prepare_fields(self, stock_pool: str):
        if stock_pool == 'all_A':
            fields = pd.DataFrame(1, index=self.load_data.index, columns=['all_A'])

        elif stock_pool == 'user_defined':
            if self.settings['data_sources'] == 'mysql':
                fileds_sql = """
                SELECT tradedate, wind_code, ifin_selfdefined_stockpool FROM dailydata WHERE tradedate BETWEEN %s AND %s
                """ % (str(self.settings['start_date']), str(self.settings['end_date']))
                fields = self.read_sql(fileds_sql)
            elif self.settings['data_sources'] == 'h5':
                fields = self.read_h5(h5file='selfdefine_stockpool.h5', key='dailydata', columns=['ifin_selfdefined_stockpool'])

        elif stock_pool in self.bm_name_code.keys():
            if self.settings['data_sources'] == 'mysql':
                fields_sql = """
                SELECT tradedate, component_code, weight FROM index_weight_daily WHERE (index_code = '%s') AND tradedate BETWEEN %s AND %s
                """ % (self.bm_name_code[stock_pool], str(self.settings['start_date']), str(self.settings['end_date']))
                fields = self.read_sql(fields_sql).rename(columns={'component_code': 'wind_code'}).set_index(['tradedate', 'wind_code'])
            elif self.settings['data_sources'] == 'h5':
                cond = """
                index_code = '%s'&tradedate>='%s'&tradedate<='%s'
                """ % (self.bm_name_code[stock_pool], str(self.settings['start_date']), str(self.settings['end_date']))
                fields = self.read_h5(h5file='index_weight_daily.h5', key='index_weight', cond=cond).rename(columns={
                    'component_code': 'wind_code'
                }).set_index(['tradedate', 'wind_code'])['weight']

        fields = fields.astype(bool).sort_index()
        fields.columns = [stock_pool]
        return fields

    @utils.func_timer
    def prepare_data(self):

        if self.settings['data_sources'] == 'mysql':
            self.load_data_sql()
        elif self.settings['data_sources'] == 'h5':
            self.load_data_h5()
        print('load data done!')

        self.base_data['ind'] = self.load_data['firstindustrycode'].astype('int64', errors='ignore')
        ind_col = [''.join(['ind', str(i)]) for i in self.inds_name_code.keys()]

        self.ind_dummy = pd.get_dummies(self.base_data['ind'], sparse=True)
        self.ind_dummy.columns = ind_col

        self.base_data['cap'] = np.log(self.load_data['afloats'].mul(self.load_data['closeprice'])).to_frame()
        self.base_data['size'] = self.base_data['cap'].groupby('tradedate').apply(lambda x: pd.qcut(x, 3, labels=['small', 'medium', 'big']))
        self.base_data['weight'] = self.base_data['cap']
        self.base_data['close'] = self.load_data['adjclose'][:]
        del self.load_data
        print('prepare data done!')


class BaseAlphaSet(Base):
    def __init__(self, settings, ifload: bool = True) -> None:
        super().__init__(settings)
        if ifload:
            self.load_alpha()

    @utils.func_timer
    def load_alpha(self):
        if self.settings['data_sources'] == 'mysql':
            alpha_sqlexp = """
            SELECT trade_date as tradedate, security_id as wind_code, %s FROM %s WHERE trade_date BETWEEN %s AND %s
            """ % (', '.join(
                self.settings['alpha_factor_list']), self.settings['alpha_factor_source'], self.settings['start_date'], self.settings['end_date'])
            try:
                self.alpha = self.read_sql(alpha_sqlexp, con=self.cr_or)
            except:
                self.alpha = self.read_sql(alpha_sqlexp, con=self.conn)

        elif self.settings['data_sources'] == 'h5':
            self.alpha = self.read_h5(h5file=self.settings['alpha_factor_source'], key='factor_value', columns=self.settings['alpha_factor_list'])

        self.close()

    def pre_process(self,
                    dataset: BaseDataSet,
                    mode: str = 'mad',
                    multiple: int = 5,
                    method: str = 'ffill',
                    ifmp: bool = False,
                    w_method: str = 'cap',
                    freq: Union[str, None] = None):
        pass


class BaseFactorCheck:
    def __init__(self, dataset: BaseDataSet, alphaset: BaseAlphaSet) -> None:
        self.dataset = dataset
        self.factor_data = utils.get_factor_data(alphaset.alpha, dataset.base_data)
        self.alpha_names = utils.get_alpha_names(self.factor_data.columns)

        # 回测结果
        self.IC = {}
        self.group_return = {}
        self.long_short = {}
        self.factor_return = {}

        # 回测信息
        self.factor_check_info = {}
        self.check_info()

        # 单因子回测结果：
        self.image = {}
        self.table = {}

    def check_info(self):
        self.factor_check_info = {
            '因子名称': None,
            '全市场域': self.dataset.fields.index[self.dataset.fields['all_A']].get_level_values(1).nunique(),
            '自定义域': self.dataset.fields.index[self.dataset.fields['user_defined']].get_level_values(1).nunique(),
            '回测起始': self.factor_data.index.levels[0][0].strftime('%Y-%m-%d'),
            '回测截至': self.factor_data.index.levels[0][-1].strftime('%Y-%m-%d'),
            '回测因子': None,
            '对照因子': None,
            '回测耗时': None,
        }

    @utils.func_timer
    def run(self):
        self.IC_analysis()
        self.factor_group_analysis()
        self.long_short_analysis()
        self.regression_analysis()
        [self.factor_report(alpha_name) for alpha_name in self.alpha_names]

    def factor_report(self, alpha_name):
        self.get_table(alpha_name)
        self.plot(alpha_name)
        self.render(alpha_name)

    def get_table(self, alpha_name):
        self.factor_check_info['因子名称'] = alpha_name
        self.table.update({'info': pd.DataFrame(self.factor_check_info, index=['值']).T})

        cols: List[str]  # = func(alpha_name, self.factor_data.columns), 其中: func: Callable #*Note 以下三条语句的cols可能是不同的.
        reg: str  # = one of regs (self.factor_return.keys())
        ic: str  # = one of ICs (self.IC.keys())
        group: str  # = one of groups (self.group_return.keys())
        self.table.update({'stats': utils.factor_return_stats(self.factor_data[cols], self.factor_return[reg][cols])})
        self.table.update({'IC': utils.print_IC_table(self.IC[ic][cols])})
        self.table.update({'groups': utils.group_return_stats(self.group_return[group][cols])})
        ...

    def plot(self, alpha_name: str):

        self.image.clear()

        name_show: str
        cols: List[str]  # = func(alpha_name, self.factor_data.columns)  #*Note 以下语句的cols可能是不同的.
        reg: str  # = one of regs (self.factor_return.keys())
        ic: str  # = one of ICs (self.IC.keys())
        group: str  # = one of groups (self.group_return.keys())

        # image1
        data = self.IC[ic][cols]
        self.image.update(dict(IC=utils.plot_line(data, ylabel='IC累计值', method='sum', suptitle='全市场股票池IC累计值', alpha_name=alpha_name)))

        # image2
        data = self.factor_return[reg].xs('beta', level=1).loc[:, cols]
        self.image.update(dict(reg=utils.plot_line(data, test='fmtest', suptitle='回归系数累计值', alpha_name=alpha_name)))

        # image3
        data = self.group_return[group].loc[:, cols].squeeze().unstack(level='fac_qt')
        self.image.update(dict(groups2=utils.plot_line(data, test='mttest', suptitle='%s行业分层分组测试' % name_show, alpha_name=alpha_name)))

        ...

    def render(self,
               alpha_name: str,
               template_path: str = 'template.xlsx',
               save_path: str = 'factor_report',
               imgsize: Tuple = (602.0016, 287.3881)) -> None:

        out_path = '{}.xlsx'.format(alpha_name)
        if save_path is not None:
            out_path = os.path.join(save_path, out_path)
            if not os.path.exists(save_path):
                os.makedirs(save_path)
        wb = openpyxl.load_workbook(template_path)
        tab_pos, img_pos = utils.get_pos(wb)
        for img_name, img in self.image.items():
            sheet_name, row, col = img_pos[img_name]
            img = Image(img)
            img.width, img.height = imgsize
            wb[sheet_name].add_image(img, anchor=get_column_letter(col) + str(row))
        for tab_name, tab in self.table.items():
            sheet_name, row, col = tab_pos[tab_name]
            utils.excel_writer(tab, wb[sheet_name], row, col)
        wb.save(out_path)

    @utils.func_timer
    def IC_analysis(self):
        # for example
        ICs: list = ['ic1', 'ic2', 'ic3']
        self.IC.update(ic1=utils.factor_information_coefficient(self.factor_data))
        self.IC.update(ic2=utils.factor_information_coefficient(self.factor_data, method='pearson'))
        self.IC.update(ic2=utils.factor_information_coefficient(self.factor_data, method='pearson', detail=True))
        ...

    @utils.func_timer
    def factor_group_analysis(self):
        # for example
        alpha_cols: list = ['alpha1', 'alpha2', ...]
        self.group_return.update(dict(group1=utils.group_return(self.factor_data[alpha_cols])))
        self.group_return.update(dict(group2=utils.group_return(self.factor_data[alpha_cols])))
        ...

    @utils.func_timer
    def long_short_analysis(self):
        # fro example
        groups1: list = ['group1', 'group2', ...]  # 在factor_group_analysis里的分组
        groups2: list = ['group3', 'group4', ...]  # 在factor_group_analysis里的分组
        self.long_short.update(dict(long_short1=pd.concat([utils.long_short_return(self.group_return(g), g) for g in groups1], axis=1)))
        self.long_short.update(dict(long_short2=pd.concat([utils.long_short_return(self.group_return(g), g) for g in groups2], axis=1)))
        ...

    @utils.func_timer
    def regression_analysis(self):
        regs: list = ['reg1', 'reg2', ...]
        self.factor_return.update(reg1=utils.factor_return(self.factor_data))


if __name__ == '__main__':
    dataset = BaseDataSet(settings)
    alphaset = BaseAlphaSet(settings)
    alphaset.pre_process(dataset)
    factorcheck = BaseFactorCheck(dataset, alphaset)
    factorcheck.run()
